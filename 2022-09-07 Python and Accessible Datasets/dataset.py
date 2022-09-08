# -*- coding: utf-8 -*-
"""
Created on Tue Mar 15 11:27:27 2022
@author: osmonj
"""

# How will guidance sheet, contents sheet sources work?


import xlsxwriter
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

 
def produce_dataset(write_path, init_template_path, dataframes,
                    font='Arial', data_align = 'right', data_font_size = 12):
    """
    Writes a dataframe in the desired format to Excel.
 
    Parameters
    ----------
    write_path : string
        File path for the dataset to be written to.
    init_table_path : string
        File path to the initialisation template, which should be filled out
        appropriately.
    dataframes : list
        List of dataframes to be written into the dataset
    font : string, optional
        Font of the dataset. The default is 'Arial'.
    data_align: string, optional
        Determines whether data and corresponding column headings in data
        tables should be aligned to the 'left', 'center', or 'right'.
        Best practice is to align to the right. The default is 'right'.
    data_font_size: int, optional
        Determines the font size in data tables. Best practice is 12 or over.
        The default is 12.
 
    Returns
    -------
    None.
 
    """
   
       
    
    ##################### Read in initialisation template #####################
   
    init = pd.ExcelFile(init_template_path)
    init_sheets = init.sheet_names
    init.close()
   
    
    index_sheets, contents_sheets = [], []
    for sheet_name in init_sheets:
        if sheet_name[0:5] == 'Index':
            index_sheets.append(sheet_name)
        if sheet_name[0:8] == 'Contents':
            contents_sheets.append(sheet_name)
    n_index_sheets = len(index_sheets)
    n_contents_tables = len(contents_sheets)
   
    notes_template = pd.read_excel(init_template_path, 'Notes', header = 0)
    number_of_notes = notes_template.shape[0]
   
    n_dataframes = len(dataframes)
   
    workbook = xlsxwriter.Workbook(write_path)
   
    ################################## Fonts ##################################
   
    # All fonts used in the dataset are defined here, apart from the format for
    # numbers, number_format, which is defined later on.
   
    basic_format_properties = {'bold': False, 'font': font,
                               'font_size': 12, 'valign': 'vcenter'}
       
    basic_format = workbook.add_format(basic_format_properties)
   
    wrap_text_format = workbook.add_format(basic_format_properties)
    wrap_text_format.set_text_wrap(True)
    wrap_text_format.set_align('left')
   
    heading_1_format = workbook.add_format({'bold': True,
                                            'font' : font,
                                            'font_size': 16,
                                            'valign': 'vcenter'})
   
    guidance_table_properties = {'bold': True, 'font' : font,
                                 'font_size': 13, 'align': 'center',
                                 'valign': 'vcenter', 'text_wrap': True}
   
    
    guidance_table_format = workbook.add_format(guidance_table_properties)
   
    guidance_table_2_format = workbook.add_format(guidance_table_properties)
    guidance_table_2_format.set_font_size(12)
   
    
    cover_heading_2_format = workbook.add_format({'bold': True,
                                                  'font' : font,
                                                  'font_size': 13,
                                                  'valign': 'vcenter'})
   
    index_format = workbook.add_format({'bold': True, 'font': font,
                                        'font_size': data_font_size,
                                        'align': 'left',
                                        'valign': 'vcenter'})
   
    table_heading_format = workbook.add_format({'bold': True, 'font' : font,
                                                'font_size': data_font_size+1,
                                                'align': data_align,
                                                'valign': 'vcenter',
                                                'text_wrap': True})
   
    table_subheading_format = workbook.add_format({'bold': True,
                                                   'font': font,
                                                   'font_size': data_font_size,
                                                   'align': data_align,
                                                   'valign': 'vcenter'})
   
    link_format = workbook.add_format()
    link_format = workbook.get_default_url_format() 
    link_format.set_align('vcenter')
    link_format.set_text_wrap(True)
    link_format.set_font(font)
    link_format.set_size(12)
   
    
    ############################### Cover sheet ###############################

    cover_sheet_template = pd.read_excel(init_template_path, 'Cover_sheet',
                                         header = None)
    cover_sheet_template = cover_sheet_template.set_index(0)
    wb_title = cover_sheet_template.loc['Title', 1]
    wb_summary = cover_sheet_template.loc['Summary', 1]
    url = cover_sheet_template.loc['URL', 1]
    url_text = cover_sheet_template.loc['URL text (optional)', 1]
    publication_date = cover_sheet_template.loc['Publication date', 1]
    next_release = cover_sheet_template.loc['Next release date', 1]
    email = cover_sheet_template.loc['Email', 1]
    phone = cover_sheet_template.loc['Telephone', 1]
   
    url_text = url_text if not pd.isnull(url_text) else wb_title
   
    cover_sheet = workbook.add_worksheet('Cover_sheet')
    cover_sheet.set_column(0, 0, 95.29)
    cover_sheet.write(0, 0, wb_title, heading_1_format)
    cover_sheet.write(1, 0, wb_summary, wrap_text_format)
    cover_sheet.write(2, 0, url, link_format, url_text)
    cover_sheet.write(3, 0, 'Publication date', cover_heading_2_format)
    cover_sheet.write(4, 0, publication_date, basic_format)
    cover_sheet.write(5, 0, 'Next release', cover_heading_2_format)
    cover_sheet.write(6, 0, next_release, basic_format)
    cover_sheet.write(7, 0, 'Contact details', cover_heading_2_format)
    cover_sheet.write(8, 0, f'mailto:{email}', link_format, email)
    cover_sheet.write(9, 0, f'Telephone: {phone}', basic_format)

    changes_template = pd.read_excel(init_template_path, 'Changes',
                                     header = 0)
   
    changes = changes_template['Changes and notes'].values
   
    if changes.shape[0] != 0:
        cover_sheet.write(10, 0, 'Changes and notes',
                          cover_heading_2_format)
        for i, change in enumerate(changes):
            cover_sheet.write(11+i, 0, change, wrap_text_format)
       
        
        
    ################################# Contents ################################
   
    # if there is more than one section that dataframes can be sorted into,
    # e.g. Annual and Quarterly, then Guidance tables will be added to the
    # contents in their own separate table. Otherwise, Guidance tables will
    # be put into the same contents table as the data tables.
    n_contents_tables = n_contents_tables\
        if n_contents_tables == 1 else n_contents_tables + 1
   
    contents = workbook.add_worksheet('Table_of_contents')
    contents.set_column(0, 0, 28)
    contents.set_column(1, 1, 56)
    contents.set_column(2, 2, 25)
    contents.set_column(3, 3, 66)
    contents.set_column(4, 4, 30)
    contents.write(0, 0, 'Table of contents', heading_1_format)
    # There may be more than one table, so this note writes an appropriate
    # note saying how many.
    contents.write(1, 0, number_of_tables_note(n_contents_tables),
                   basic_format)
   
    # There will always be empty Estimate and Units tables for the Guidance
    # table, so a note in cell A3 is always necessary.
    contents.write(2, 0,
                   'Some cells in the Estimate, Units and Sources columns '
                   'below have been left empty where there are no '
                   'estimates or references.',
                   basic_format)
   
    n_guidance_sheets = 1+n_index_sheets+(1 if number_of_notes>0 else 0)
   
    
    # we want the correct number of Sources columns
    n_sources_cols = 1
   
    ## This next bit is probably inefficient
   
    list_of_contents_sections = []
    list_of_contents_tables = []
   
    for sheet_name in contents_sheets:
        contents_template = pd.read_excel(init_template_path, sheet_name,
                                          header = None)
        columns = contents_template.iloc[2,:].values
       
        contents_table = \
            contents_template.copy().iloc[3:,:].reset_index(drop=True)
        contents_table = contents_table.rename(
            columns = dict(zip(contents_table.columns, columns))
            )
       
        contents_sections = contents_template.copy().iloc[:2,:2]
        contents_sections = contents_sections.set_index(0).T
        list_of_contents_sections.append(contents_sections)
        list_of_contents_tables.append(contents_table)
       
        # This checks whether this table has more 'Sources' columns than
        # already accounted for
        sources = contents_table.filter(regex = r'Sources_', axis = 1)
       
        last_source_col = sources.columns[-1]
        last_source_col_number = int(last_source_col[8:])
        if last_source_col_number > n_sources_cols:
            n_sources_cols = last_source_col_number
   
    
    sections = []
    subsections = []
    table_names = []
    for table in list_of_contents_sections:
        section = table.loc[1, 'Section name']
        subsection = table.loc[1, 'Subsection name']
       
        # Section is an optional parameter
        if section not in sections and not pd.isnull(section):
            sections.append(section)
            subsections.append([section, subsection])
        elif section in sections:
            i = sections.index(section)
            subsections[i].append(subsection)
           
        # table_names is only used if there are no sections, only
        # subsections.
        table_names.append(subsection)
           

    n_sections = len(sections)
   
    
    # headings_rows is going to tell us which rows to AVOID when filling in
    # the contents tables later
    headings_rows = []
   
    
    # The first column is missing here - it is added later
    columns_dict = [{'header': 'Estimate'},
                    {'header': 'Units'},
                    {'header': 'Table description and information'}]
   
    if n_sources_cols == 1:
        columns_dict += [{'header': 'Sources'}]
    else:
        for i in range(n_sources_cols):
            columns_dict += [{'header': f'Sources ({i+1})'}]
            # set the correct column widths for the new sources columns
            contents.set_column(4+i, 4+i, 30)                  
    
    if n_contents_tables == 1:
        # If the initialisation template only has one contents table, this
        # table will be concatenated into one table with the Guidance
        # sheets' contents table.
        columns_dict1 = [{'header': 'Sheet name'}] + columns_dict
        contents.add_table(3, 0, 3+n_guidance_sheets+n_dataframes,
                           3+n_sources_cols,
                           options = {'autofilter': False,
                                      'style': 'None',
                                      'columns': columns_dict1,
                                      'name': 'Table_of_contents'})
        contents.set_row(3, 16.8, guidance_table_format)
        contents_row = 3+n_guidance_sheets+n_dataframes+1
       
    else:
        columns_dict1 = [{'header': 'Guidance sheets'}] + columns_dict
        contents.add_table(3, 0, 3+n_guidance_sheets, 3+n_sources_cols,
                           options = {'autofilter': False,
                                      'style': 'None',
                                      'columns': columns_dict1,
                                      'name': 'Table_of_contents_1'})

        contents.set_row(3,
                         cell_format = guidance_table_format)
        contents_row = 3+n_guidance_sheets+1
   
    # The first table heading is in row 3, so it is added to heading_rows
    headings_rows.append(3)
   
    
    if n_sections == 0:
        for i in range(n_contents_tables-1):
            contents_table = list_of_contents_tables[i]
            nrows = contents_table.shape[0]
            columns_dict1 = [{'header': table_names[i]}] + columns_dict
            contents.add_table(contents_row, 0,
                               contents_row + nrows, 3+n_sources_cols,
                               options =
                               {'autofilter': False,
                                'style': 'None',
                                'columns': columns_dict1,
                                'name': f'Table_of_contents_{i+2}'})
            contents.set_row(contents_row,
                             cell_format = guidance_table_format)
            headings_rows.append(contents_row)
            contents_row += nrows+1
    else:
        i = 0
        for section in subsections:
            contents.write(contents_row, 0, section[0],
                           cover_heading_2_format)
            headings_rows.append(contents_row)
            contents_row += 1
            for subsection in section[1:]:
                contents_table = list_of_contents_tables[i]
                nrows = contents_table.shape[0]
                columns_dict1 = [{'header': table_names[i]}] + columns_dict
                contents.add_table(contents_row, 0,
                                   contents_row + nrows, 3+n_sources_cols,
                                   options =
                                   {'autofilter': False, 'style': 'None',
                                    'columns': columns_dict1,
                                    'name': f'Table_of_contents_{i+2}'})
                contents.set_row(contents_row,
                                 cell_format = guidance_table_2_format)
                headings_rows.append(contents_row)
                contents_row += nrows+1
                i += 1
           
    

    

    # This variable will increase every time a new sheet is added to the
    # workbook, so that the next sheet is added in the next row of the
    # contents table.
    non_headings_rows = []
    # This range takes us to the last row in the contents sheet
    for i in range(4, contents_row+1):
        if i not in headings_rows:
            non_headings_rows.append(i)
    non_headings_rows = iter(non_headings_rows)
    contents_row = next(non_headings_rows)
   
    ################################# Guidance ################################
   
    guidance_template = pd.read_excel(init_template_path, 'Guidance',
                                      header = 0)
   
    n, m = guidance_template.shape
   
    # Count number of reference columns
    n_refs = 0
    for col in guidance_template.columns:
        if col[:9] == 'Reference':
            n_refs += 1
           
    guidance = workbook.add_worksheet('Guidance')
    guidance.set_column(0, 0, 28)
    guidance.set_column(1, 1, 120)
    
    guidance.write(0, 0, 'Sources and methodology', heading_1_format)
    guidance.write(1, 0, 'This worksheet contains one table.',
                   basic_format)
   
    # Check to see if the last references column has any missing data. If
    # so, a note at the top of the worksheet explaining this is necessary.
    if sum(pd.isnull(guidance_template)[f'Reference_{n_refs}']) > 0:
        first_table_row = 3
        guidance.write(2, 0,
                       'Some cells in the \'Sources and references\' '
                       'columns have been left blank because there were '
                       'no relevant sources.', basic_format)
    else:
        first_table_row = 2
   
        
    
    guidance_cols_dict = [{'header': 'Notes'}, {'header': 'Guidance'}]
    if n_refs == 1:
        guidance_cols_dict.append({'header': 'Sources and references'})
        guidance.set_column(2, 2, 34.44)
    elif n_refs == 0:
        pass
    else:
        for i in range(1, n_refs+1):
            guidance_cols_dict.append({'header':
                                       f'Sources and references ({i})'})
            guidance.set_column(1+i, 1+i, 34.44)
    
    # Want to ignore any URL columns, of which there must be as many as
    # there are Reference columns, so exclude those from table.
    guidance.add_table(first_table_row, 0, first_table_row+n, m-1-n_refs,
                       options = {'autofilter': False,
                                  'style': 'None',
                                  'columns': guidance_cols_dict,
                                  'name': 'Guidance'})
    guidance.set_row(first_table_row, 16.8, guidance_table_format)
   
    for i in range(n):
        for j in range(2):
            try:
                guidance.write(first_table_row+1+i, j,
                               guidance_template.iloc[i,j],
                               wrap_text_format)
            except:
                pass
        for j in range(1, n_refs+1):
            try:
                if pd.isnull(guidance_template.loc[i, f'URL_{j}']):
                    guidance.write(first_table_row+1+i, 1+j,
                                   guidance_template.loc[i,
                                                         f'Reference_{j}'],
                                   wrap_text_format)
                else:
                    guidance.write(first_table_row+1+i, 1+j,
                                   guidance_template.loc[i, f'URL_{j}'],
                                   link_format,
                                   guidance_template.loc[i,
                                                         f'Reference_{j}'])
            except:
                pass
   
    contents.write(contents_row, 0, 'internal:Guidance!A1', link_format,
                   'Guidance')
    # No estimates or units
    contents.write(contents_row, 3,
                   'This sheet contains notes on dataset structure and '
                   'methodology, and additional notes.',
                   wrap_text_format)
    # No sources or references
   
    
    ################################## Index ##################################
   
    for sheet_name in index_sheets:
       
        index_template = pd.read_excel(init_template_path, sheet_name,
                                       header = None)
       
        
        if index_template.shape[0] <= 5:#.iloc[5,0] in [None, nan]:
            # stops the creation of more index sheets if they're empty
            # below the fifth cell
            break
       
        # Row 5 (6 in Excel) contains the header row of the index table
        columns = index_template.iloc[5,:].values
        index_template = index_template.set_index(0)
       
        index_tab_name = index_template.loc['Tab name', 1]
        index_title = index_template.loc['Sheet title', 1]
        index_description = index_template.loc['Contents description', 1]
        index_source = index_template.loc['Source', 1]
        index_source_url = index_template.loc['Source URL', 1]
           
        index_template = index_template.rename(
            columns = dict(zip(index_template.columns,
                               index_template.iloc[5,:].values))
            )
        index_template = index_template.iloc[6:,:].reset_index()
       
        n, m = index_template.shape
       
        index = workbook.add_worksheet(index_tab_name)
       
        for j in range(m):       
            index.set_column(j, j, 50)
        index.write(0, 0, index_title, heading_1_format)
        index.write(1, 0, 'This worksheet contains one table.',
                    basic_format)
       
        
        table_cols_dict = []
       
        for col in columns:
            table_cols_dict.append({'header': col})
        index.add_table(2, 0, 2+n, m-1,
                        options = {'autofilter': False,
                                   'style': 'None',
                                   'columns': table_cols_dict,
                                   'name': index_tab_name})
        index.set_row(2, 16.8, guidance_table_format)
       
        for i in range(n):
            for j in range(m):
                try:
                    index.write(3+i, j, index_template.iloc[i,j],
                                wrap_text_format)
                except:
                    pass
        contents_row = next(non_headings_rows)
        contents.write(contents_row, 0, f'internal:{index_tab_name}!A1',
                       link_format, index_tab_name)
        # No estimates or units
        contents.write(contents_row, 3, index_description,
                       wrap_text_format)
       
        if not pd.isnull(index_source):
            if pd.isnull(index_source_url):
                contents.write(contents_row, 4, index_source,
                               wrap_text_format)
            else:
                contents.write(contents_row, 4, index_source_url,
                               link_format, index_source)
       
        
            
    
    ################################## Notes ##################################
   
    if number_of_notes != 0:
        notes = workbook.add_worksheet('Notes')
        notes.set_column(0, 0, 28)
        notes.set_column(1, 1, 104)
       
        notes.write(0, 0, 'Notes', heading_1_format)
        notes.write(1, 0, 'This worksheet contains one table.',
                    basic_format)
       
        notes.add_table(2, 0, 2+number_of_notes, 1,
                        options = {'autofilter': False,
                                   'style': 'None',
                                   'columns': [{'header': 'Number'},
                                               {'header': 'Note'}],
                                   'name': 'Notes'})
        notes.set_row(2, 16.8, guidance_table_format)
       
        notes_descriptions = notes_template['Note'].values
       
        for i, description in enumerate(notes_descriptions, start = 1):
            notes.write(2 + i, 0, f'Note {i}',
                        basic_format)
            notes.write(2 + i, 1, description, wrap_text_format)
        contents_row = next(non_headings_rows)
        contents.write(contents_row, 0, 'internal:Notes!A1', link_format,
                       'Notes')
        # No estimates or units
        contents.write(contents_row, 3,
                       'This sheet contains additional notes specific to '
                      'aspects of various tables in this workbook.',
                       wrap_text_format)
        # No sources or references
       
        
    ################################## Tables #################################
   
    dfs_metadata = pd.concat(list_of_contents_tables, axis=0)
    dfs_metadata = dfs_metadata.reset_index(drop=True)
   
    for x in range(dfs_metadata.shape[0]):
       
        df = dataframes[x]
       
        md = dfs_metadata.loc[x, :]
       
        hidden = md['Hidden decimal places']
        if not pd.isnull(hidden):
            # Some values might be text, so we need to avoid this.
            # There should be a more efficient method for this.
            for i in range(df.shape[0]):
                for j in range(df.shape[1]):
                    try:
                        df.iloc[i,j] = round(df.iloc[i, j], hidden)
                    except TypeError:
                        # This picks up string values
                        pass
            #df = df.round(hidden)
           

    
        display = md['Displayed decimal places']
        display = 1 if pd.isnull(display) else display
       
        # Creates a string with as many zeroes as desired decimal
        # places, for use in the f-string deciding number formats
        zeros = '0' * display 
        num_format = f'#,##0.{zeros}' if display > 0 else '#,##0'
        number_format = workbook.add_format({'num_format': num_format,
                                             'bold': False,
                                             'font' : font,
                                             'font_size': data_font_size,
                                             'align': data_align,
                                             'valign': 'vcenter'})
       
        
        
        table_num = f'Table_{x+1}'
       
        worksheet = workbook.add_worksheet(table_num)
        worksheet.set_column(0, df.shape[1], 28)
       
        
        worksheet.write(0, 0, 'Worksheet: ' + md['Name'], heading_1_format)
        worksheet.write(1, 0, 'This worksheet contains one table.',
                        basic_format)
       
        
        missing_data_note = md['Missing data note']
        # This if statement checks for missing data and attaches a note
        # if any data are missing
        if pd.isnull(df).to_numpy().sum() > 0 or\
            not pd.isnull(missing_data_note):

            cols_with_na_data = []
           
            for column in df.columns:
                if pd.isnull(df[column]).sum() > 0:
                    cols_with_na_data.append(column)
            plural = 's are' if len(cols_with_na_data) > 1 else ' is'
           
            missing_data_note = \
                f'The {list_to_sentence(cols_with_na_data)} '\
                f'column{plural} missing data.' \
                if pd.isnull(missing_data_note) else missing_data_note
           
            worksheet.write(2, 0, missing_data_note, basic_format)
            extra_row = 1
        else:
            extra_row = 0
           
        column_dict_list = [{'header': df.index.name}]
        for col_num, col in enumerate(df.columns):
           
            column_dict_list.append({'header': col})   
        
        
        table_markup_name = md['Marked up table name']
        table_markup_name = table_num if pd.isnull(table_markup_name) \
            else table_markup_name
       
        worksheet.add_table(2+extra_row, 0,
                            2+extra_row+df.shape[0], df.shape[1],
                            options = {'autofilter': False,
                                        'style': 'None',
                                        'columns': column_dict_list,
                                        'name': table_markup_name})
        worksheet.set_row(2+extra_row, 90, table_heading_format)
       
        
        n_headers = md['Number of header rows']
        n_headers = 1 if pd.isnull(n_headers) else n_headers
       
        for i in range(df.shape[0]):
            worksheet.write(3+extra_row+i, 0, df.index[i], index_format)
       
        
        # Writes the data itself to the worksheet, in a format that rounds
        # all entries to the same number of decimal places
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                try:
                    if i < n_headers-1:
                        worksheet.write(3+extra_row+i, 1+j, df.iloc[i,j], 
                                        table_subheading_format)
                    else:
                        worksheet.write(3+extra_row+i, 1+j, df.iloc[i,j], 
                                        number_format)
                except:
                    pass
               
        contents_row = next(non_headings_rows)
        
        
        estimate = '' if pd.isnull(md['Estimate']) else md['Estimate']
        unit = '' if pd.isnull(md['Unit(s)']) else md['Unit(s)']
        description = \
            '' if pd.isnull(md['Description']) else md['Description']
        contents.write(contents_row, 0, f'internal:{table_num}!A1',
                       link_format, table_num)
        contents.write(contents_row, 1, estimate, wrap_text_format)
        contents.write(contents_row, 2, unit, wrap_text_format)
        contents.write(contents_row, 3, description, wrap_text_format)
       
        for s in range(1, n_sources_cols+1):
           
            try:
                if pd.isnull(md[f'URL_{s}']):
                    contents.write(contents_row, 3+s,
                                   md[f'Sources_{s}'],
                                   wrap_text_format)
                else:
                    contents.write(contents_row, 3+s,
                                   md[f'URL_{s}'],
                                   link_format,
                                   md[f'Sources_{s}'])
            except:
                pass
      
    try:
        workbook.close()
    except xlsxwriter.exceptions.FileCreateError:
        print(f'The dataset \'{write_path}\' could not be created. Check that '
              'a file with this file path is not already open.')
    else:
       
        ############################### Openpyxl ##############################
       
        # Reload data with openpyxl to markup certain cells as Headings.
       
        wb = openpyxl.load_workbook(write_path)
       
        heading_1_font = Font(name = font, size = 16, b = True)
        heading_2_font = Font(name = font, size = 13, b = True)
        heading_3_font = Font(name = font, size = 12, b = True)
       
        for sheet in wb.worksheets:
            n = sheet.max_row
            m = sheet.max_column
           
            for i in range(1,n+1):
                for j in range(1,m+1):
                    cell = sheet[f'{get_column_letter(j)}{i}']
                   
                    # Assuming a cell is not a hyperlink, set color to be None.
                    # This makes the text colour of that cell 'Automatic'.
                    if not cell.font.underline:
                        cell.font = Font(color = None,
                                         size = cell.font.size,
                                         name = cell.font.name,
                                         b = cell.font.b)
                    if cell.value == None:
                        cell.font = Font(color = None,
                                         size = 12,
                                         name = font)
           
            
            
            # Markup cell A1 in each sheet as Headings 1
            sheet['A1'].style = 'Headline 1'
            # Correct font of cell A1
            sheet['A1'].font = heading_1_font
            # Remove border of cell A1
            sheet['A1'].border = None
           
            # This aligns the index heading of a table to the LEFT instead of
            # the right
            for row in range(3,5):
                year_cell = f'A{row}'
                if sheet[year_cell].value in ['Year', 'Quarter']:
                    sheet[year_cell].alignment = Alignment(horizontal = 'left',
                                                           vertical = 'center')
                   
        
        cover_sheet = wb['Cover_sheet']
       
        for row in range(2,15):
            cell = f'A{row}'
            if cover_sheet[cell].value in ['Publication date',
                                           'Next release',
                                           'Contact details',
                                           'Changes',
                                           'Changes and notes']:
                cover_sheet[cell].style = 'Headline 2'
                cover_sheet[cell].font = heading_2_font
                cover_sheet[cell].border = None
       
        contents_sheet = wb['Table_of_contents']
       
        
        
        for row in range(1, contents_row+1):
            cell = f'A{row}'
            # Heading 2 should be used in the contents to mark tables if and
            # only if there are no other headings
           
            cell_value = contents_sheet[cell].value
           
            if cell_value in ['Guidance sheets']+sections:
               
                contents_sheet[cell].style = 'Headline 2'
                contents_sheet[cell].font = heading_2_font
                contents_sheet[cell].alignment = Alignment(horizontal = 'left',
                                                           vertical = 'center',
                                                           wrap_text = True)
                contents_sheet[cell].border = None
            elif cell_value in table_names and n_sections == 0:
               
                contents_sheet[cell].style = 'Headline 2'
                contents_sheet[cell].font = heading_2_font
                contents_sheet[cell].alignment = Alignment(horizontal = 'left',
                                                           vertical = 'center',
                                                           wrap_text = True)
                contents_sheet[cell].border = None
            # if there are other headings
            elif cell_value in table_names:
               
                contents_sheet[cell].style = 'Headline 3'
                contents_sheet[cell].font = heading_3_font
                contents_sheet[cell].alignment = Alignment(horizontal = 'left',
                                                           vertical = 'center',
                                                           wrap_text = True)
                contents_sheet[cell].border = None
       
        
        wb.save(write_path)
 
    
def number_of_tables_note(n):
    """
    This function produces a note to add to the top of a worksheet, such as:
           
        This worksheet contains one table.
       
        This worksheet contains three tables, stacked vertically.
       
        This worksheet contains 12 tables, stacked vertically.
 
    """
   
    number2word = dict(zip(range(1,10+1),
                           ['one', 'two', 'three', 'four', 'five', 'six',
                            'seven', 'eight', 'nine', 'ten']))
    word = number2word[n] if n <= 10 else n
   
    tables_note =  f'This worksheet contains {word} table'
    tables_note += 's, stacked vertically.' if n != 1 else '.'
   
    return tables_note
 

def list_to_sentence(alist, oxford_comma = True):
    """
    Parameters
    ----------
    alist : list
        A list of objects to be joined together in a list.
    oxford_comma : bool, optional
        Decides whether or not to include an Oxford comma in the list.
        The default is True.
 
    Returns
    -------
    sentence : str
        The items in the string concatenated in sentence form.
 
    """
   
    if not isinstance(alist, list):
        alist = [alist]
    if len(alist) == 1:
        sentence = str(alist[0])
    elif len(alist) == 2:
        sentence = f'{alist[0]} and {alist[1]}'
    else:
        sentence = ', '.join(alist[:-1])
        sentence += f', and {alist[-1]}'\
            if oxford_comma else f' and {alist[-1]}'
   
    return sentence
           
 
if __name__ == '__main__':
    import numpy as np
    
    np.random.seed(42)
    df1 = pd.DataFrame(np.random.randn(2022-1980, 1)*100,
                       index = list(range(1980, 2022)),
                       columns = ['UK'])
    df1.index.name = 'Year'
    
    
    trusts = ['RX6','RX7','RX8','RX9','RYA','RYC','RYD','RYE','RYF','RRU']
    regions = ['North East and Yorkshire','North West',
               'North East and Yorkshire','Midlands', 'Midlands',
               'East of England','South East','South East','South West',
               'London']
    regions = np.array(regions)
    regions.transpose()
    
    df2 = pd.DataFrame(np.random.randn(2022-1980, 10)*100,
                       index = list(range(1980, 2022)),
                       columns = trusts)
    df2.index.name = 'Year'
    df2_region = pd.DataFrame(regions, index=trusts, columns = ['Region']).T
    df2_region.index.name = 'Year'
    df2 = pd.concat([df2_region, df2], axis=0)
    
    
    produce_dataset('my_dataset.xlsx', 'test template.xlsx', 
                    [df1, df2, df2, df2])



    
    
